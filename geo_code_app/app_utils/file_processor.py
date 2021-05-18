from openpyxl import load_workbook,Workbook
from openpyxl.writer.excel import save_virtual_workbook
import geocoder
import json
import uuid
from datetime import datetime 
import os

def process_file(file_obj):
    try:
        ex_wb = load_workbook(file_obj)
        ex_sheets = ex_wb.sheetnames
        worksheet = ex_wb[ex_sheets[0]]
        book = Workbook()
        output_sheet = book.active
        output_sheet.append(('Address','Latitude','Longitude'))
        for cell in worksheet['A']:
            if cell.value.strip() == '':
                output_sheet.append((cell.value,'Empty Address','Empty Address'))
            else:
                output_sheet.append(get_lat_lang(cell.value))
        return {"message":"success","data":save_virtual_workbook(book),"file_name":get_file_name()}

    except Exception as e:
        return {"message":"failure"}

def get_lat_lang(address):
    try:
        g =  geocoder.mapquest(address,key=os.environ["MAP_QUEST_KEY"])
        if g.status_code == 200:
            if g.status == 'OK':
                results = g.json
                lat_obj = results['raw']['latLng']
                return (address,lat_obj['lat'],lat_obj['lng'])
            else:
                return (address,g.status,g.status)
        else:
            return (address,json.dumps({"status_code":g.status_code}),json.dumps({"status_code":g.status_code}))
    except Exception as e:
        print("exception",e)
        return (address,"Exception occured","Exception occured")

def getDateTimeStr():
    t_format_string = '%Y_%m_%d_%H_%M_%S'
    return datetime.now().strftime(t_format_string)
    
def get_file_name():
    return getDateTimeStr()+str(uuid.uuid4())+".xlsx"